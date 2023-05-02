from openpyxl import Workbook, load_workbook
import os
from constant.constant import WORKSHEET_TILE


class ExportGachaRecord(object):
    def __init__(self, workdir):
        self.workdir = workdir
        os.chdir(workdir)

    def export_gacha_record(self, gacha_type, gacha_res):
        export_name = '%s.xlsx' % gacha_res[0]['uid']

        gacha_res = self.json2list(gacha_res=gacha_res)
        # print(gacha_res)
        wb = self.check_xlsx(export_name)
        if gacha_type == 1:
            # 常驻
            ws = self.check_sheet('常驻池', wb)
            self.write_data(ws, gacha_res)
        elif gacha_type == 11:
            # 角色
            ws = self.check_sheet('角色池', wb)
            self.write_data(ws, gacha_res)
        elif gacha_type == 12:
            # 光锥
            ws = self.check_sheet('光锥池', wb)
            self.write_data(ws, gacha_res)

        wb.save(export_name)

    def json2list(self, gacha_res):
        gacha_list = []
        for gacha in gacha_res:
            gacha_list.append(list(gacha.values()))

        return gacha_list

    def check_sheet(self, sheet_name, workbook):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)
            return self.init_worksheet(ws)


    def check_xlsx(self, xlsx_name):
        print(os.listdir(self.workdir))
        if xlsx_name in os.listdir(self.workdir):
            return load_workbook(xlsx_name)
        else:
            return Workbook()

    def write_data(self, worksheet, gacha_res):
        for gacha in gacha_res:
            worksheet.append(gacha)

    def init_worksheet(self, worksheet):
        worksheet.append(WORKSHEET_TILE)
        return worksheet




